from openpyxl import Workbook
import csv
input_path='c:/folder/file.txt'
output_path='c:/folder/file.xlsx'
wb=Workbook()
sheet=1
title=''
with open(input_path, 'r') as dump_file:
    csv_reader= csv.reader(dump_file,delimiter=',',quotechar='"')
    ws=None
    sheet_position=1
    for col1 in csv_reader:
      ws=wb.create_sheet(title+str(sheet),sheet)
      ws['A'+str(sheet_position)]=col1
      sheet_position+=1
wb.save(output_path)
